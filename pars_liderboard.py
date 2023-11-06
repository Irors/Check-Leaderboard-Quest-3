import aiohttp
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import openpyxl
import asyncio
from pydantic import BaseModel
import logging
from loguru import logger

questPass = '0x03e88d43a310633152deef7d164dd4273eb2ce8b0ffc0d1ff597ab49fd88908d::quest::QuestPass'
capy = '0x2::kiosk::KioskOwnerCap'

class Check_data(BaseModel):
    reward: int
    score: int
    rank: int | None
    bot: bool

class Check_apps(BaseModel):
    appsUsed: list
    IS_ELIGIBLE: bool


class Excel:

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Address'
    sheet['G1'] = 'Eligible'
    sheet['C1'] = 'AppsUsed'
    sheet['D1'] = 'Rank'
    sheet['E1'] = 'Score'
    sheet['F1'] = 'Bot'
    sheet['B1'] = 'Balance'
    sheet['H1'] = 'Has a nft'

    # Настройка размера столбца A
    column = sheet.column_dimensions[openpyxl.utils.get_column_letter(1)]
    column.width = 16  # Задаем ширину столбца (в пикселях)

    # Настройка размера столбца B
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(2)]
    column_B.width = 16

    # Настройка размера столбца C
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(3)]
    column_B.width = 18

    # Настройка размера столбца D
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(4)]
    column_B.width = 16

    # Настройка размера столбца E
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(5)]
    column_B.width = 16

    # Настройка размера столбца F
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(6)]
    column_B.width = 16

    # Настройка размера столбца G
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(7)]
    column_B.width = 20

    # Настройка размера столбца G
    column_B = sheet.column_dimensions[openpyxl.utils.get_column_letter(8)]
    column_B.width = 16


    """ Декорации Excle """
    # Создаем объект для задания стиля выравнивания и отступов
    alignment_style = Alignment(horizontal='center', vertical='center', indent=3, wrap_text=True)

    # Создаем объект для задания стиля границ
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


    # Делаем все ячейки с жирным шрифтом
    for i in ('ABCDEFGH'):
        cell = sheet[f'{i}1']
        bold_font = Font(bold=True)
        cell.font = bold_font

        # Применяем стиль выравнивания и отступов к ячейке
        cell.alignment = alignment_style
        # Применяем стиль границ
        cell.border = border_style



async def fetch(number, address, params, session):
    async with session.get('https://quests.mystenlabs.com/api/trpc/user', params=params) as response:
        response = await response.json()

        if response[0]['result']['data'] is None:
            Excel.sheet[f'A{number}'] = address
        else:
            account_info = Check_data.parse_obj(response[0]['result']['data'])
            account_apps = Check_apps.parse_obj(response[0]['result']['data']['metadata'])

            Excel.sheet[f'A{number}'] = address
            Excel.sheet[f'G{number}'] = account_apps.IS_ELIGIBLE
            Excel.sheet[f'C{number}'] = len(account_apps.appsUsed)
            Excel.sheet[f'D{number}'] = account_info.rank
            Excel.sheet[f'E{number}'] = account_info.score
            Excel.sheet[f'F{number}'] = account_info.bot

    json_data_balance = {
        'jsonrpc': '2.0',
        'id': 1,
        'method': 'suix_getBalance',
        'params': {
            'owner': f'{address}',
        },
    }
    flag = False
    async with session.post('https://suiscan.xyz/api/sui/mainnet/', json=json_data_balance) as response:
        try:
            res_balance = await response.json()
            Excel.sheet[f'B{number}'] = int(res_balance['result']['totalBalance']) / 10**9
        except:
            pass

    if not flag:
        headers_nft = {'authority': 'sui-mainnet.blockvision.org', 'accept': '*/*','accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7', 'client-sdk-type': 'typescript','client-sdk-version': '0.41.0', 'client-target-api-version': '1.8.0','content-type': 'application/json', 'origin': 'https://portfolio.martianwallet.xyz','referer': 'https://portfolio.martianwallet.xyz/','sec-ch-ua': '"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"','sec-ch-ua-mobile': '?0', 'sec-ch-ua-platform': '"Windows"', 'sec-fetch-dest': 'empty','sec-fetch-mode': 'cors', 'sec-fetch-site': 'cross-site','user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36', }
        json_data_nft = {'jsonrpc': '2.0', 'id': '0', 'method': 'suix_getOwnedObjects', 'params': [f'{address}', {'filter': {'MatchNone': [{'StructType': '0x2::coin::Coin', }, ], },'options': {'showType': True, 'showDisplay': True, 'showContent': True, 'showBcs': False,'showOwner': False, 'showPreviousTransaction': False, 'showStorageRebate': False, }, }, None,
                                                                                                   None, ], }
        async with session.post('https://suiscan.xyz/api/sui/mainnet/', headers=headers_nft, json=json_data_nft) as response:
                response_nft = await response.json()
                if questPass in str(response_nft) or capy in str(response_nft):
                    Excel.sheet[f'H{number}'] = True
                    flag = True

    Excel.workbook.save('check.xlsx')


async def get_top(number: int, address: str):
    async with aiohttp.ClientSession() as session:
        params = {
            'batch': '1',
            'input': '{"0":{"address":"%s","questId":3}}' % address,
        }

        await fetch(number, address, params, session)


if __name__ == '__main__':

    async def main_(address_list: list, number: int = 2):
        tasks = []
        for wallet in address_list:
            tasks.append(asyncio.create_task(get_top(number, wallet)))
            number += 1

        await asyncio.gather(*tasks)


    Excel()
    #logging.getLogger().setLevel(logging.DEBUG)
    logger.info('- Start pars. -')
    with open("wallets.txt") as file:
        wallets = [i.strip() for i in file]

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    loop.run_until_complete(main_(wallets))
    logger.info('- Pars end. -')
